import json
import bs4
import requests
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
import openpyxl

with open("lineups.json") as lineups_json:
    lineups = json.load(lineups_json)


def get_text_no_recursive(parent):
    return ''.join(parent.find_all(text=True, recursive=False)).strip()


def scrap_season(year, url):
    season_data = {}

    chrome_driver = webdriver.Chrome("C:\\Users\\pcost\\chromedriver_win32\\chromedriver.exe")
    delay = 3  # delay to load page
    chrome_driver.get(url)
    try:
        myElem = WebDriverWait(chrome_driver, delay).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'matchFixtureContainer')))
        print("Page is ready!")
    except TimeoutException:
        print("Loading page took too much time!")

    soup = bs4.BeautifulSoup(chrome_driver.page_source, "html.parser")
    chrome_driver.close()

    season_data["Year"] = year
    matches_soup_list = soup.find_all("li", class_="matchFixtureContainer")
    matches_url_list = ["https://www.premierleague.com/match/" + match["data-comp-match-item"] for match in matches_soup_list]

    matches_data = [scrap_match_page(match_url) for match_url in matches_url_list]

    return {"Year": year, "Matches": matches_data}


def scrap_match_page(url):
    print(url)
    match_data = {}
    request_page = requests.get(url)
    soup = bs4.BeautifulSoup(request_page.content, "html.parser")

    match_data["Home Team"] = soup.find("div", class_="team home").find("span", class_="long").get_text()
    match_data["Away Team"] = soup.find("div", class_="team away").find("span", class_="long").get_text()
    teams_formation_soup_list = soup.find_all("strong", class_="matchTeamFormation")
    home_team_formation = teams_formation_soup_list[0].get_text()
    away_team_formation = teams_formation_soup_list[1].get_text()

    # players = {}
    home_players_soup_list = soup.find("div", class_="col-4-m").find_all("li", class_="player")
    away_players_soup_list = soup.find("div", class_="col-4-m right").find_all("li", class_="player")
    home_players = {player.find("div", class_="number").get_text(): get_text_no_recursive(player.find("div", class_="name")) for player in home_players_soup_list}
    away_players = {player.find("div", class_="number").get_text(): get_text_no_recursive(player.find("div", class_="name")) for player in away_players_soup_list}

    home_rows = []
    away_rows = []

    home_rows_soup_list = soup.find("div", class_="team home pitchPositonsContainer").find_all("div", class_="row")
    for rows_soup in home_rows_soup_list:
        row = []
        row_soup_list = rows_soup.find_all("div", class_="pos")
        for row_soup in row_soup_list:
            row.append(row_soup.get_text())
        home_rows.append(row[::-1])

    away_rows_soup_list = soup.find("div", class_="team away pitchPositonsContainer").find_all("div", class_="row")
    for rows_soup in away_rows_soup_list:
        row = []
        row_soup_list = rows_soup.find_all("div", class_="pos")
        for row_soup in row_soup_list:
            row.append(row_soup.get_text())
        away_rows.append(row[::-1])

    home_players_name = []
    away_players_name = []

    for row in home_rows:
        for player_number in row:
            home_players_name.append(home_players[player_number])

    for row in away_rows:
        for player_number in row:
            away_players_name.append(away_players[player_number])

    match_data["Home Vector"] = []
    i = 0
    for tactic_position in lineups["tactics"][home_team_formation]:
        if tactic_position == "1":
            match_data["Home Vector"].append(home_players_name[i])
            i += 1
        else:
            match_data["Home Vector"].append(tactic_position)

    match_data["Away Vector"] = []
    i = 0
    for tactic_position in lineups["tactics"][away_team_formation]:
        if tactic_position == "1":
            match_data["Away Vector"].append(away_players_name[i])
            i += 1
        else:
            match_data["Away Vector"].append(tactic_position)

    return match_data


def create_wb(seasons):

    wb = openpyxl.Workbook()
    current_ws = wb.active

    first_season = True

    for season in seasons:
        if first_season:
            first_season = False
            current_ws.title = season["Year"]
        else:
            current_ws = wb.create_sheet(season["Year"])

        current_column = 1
        current_ws.cell(row=1, column=current_column, value="HT")

        current_column += 1
        current_ws.cell(row=1, column=current_column, value="AT")

        for position in lineups["positions"]:
            current_column += 1
            current_ws.cell(row=1, column=current_column, value="H" + position)

        for position in lineups["positions"]:
            current_column += 1
            current_ws.cell(row=1, column=current_column, value="A" + position)

        current_row = 2
        current_column = 1
        for match in season["Matches"]:
            current_ws.cell(row=current_row, column=current_column, value=match["Home Team"])
            current_column += 1
            current_ws.cell(row=current_row, column=current_column, value=match["Away Team"])
            for vector_element in match["Home Vector"]:
                current_column += 1
                current_ws.cell(row=current_row, column=current_column, value=vector_element)
            for vector_element in match["Away Vector"]:
                current_column += 1
                current_ws.cell(row=current_row, column=current_column, value=vector_element)

    wb.save("lineups.xlsx")
    wb.close()


# match1 = scrap_match_page("https://www.premierleague.com/match/46975")


# create_wb([{"Year": "2018_2019", "Matches": [match1]}])

print(scrap_season("2019_2020", "https://www.premierleague.com/results?co=1&se=274&cl=-1"))
